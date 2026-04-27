#!/usr/bin/env python3
"""
Offline-friendly XLSX -> CSV converter.

- Reads an .xlsx file (all sheets).
- Exports each sheet to a separate UTF-8 (with BOM) comma-separated CSV.
- Avoids pandas/numpy to keep the exe lightweight.

Example:
  xlsx_to_csv.exe --input "your.xlsx" --out-dir ".\\out"
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


def _safe_name(name: str) -> str:
    s = str(name or "").strip()
    s = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", s)
    return (s[:120] if s else "Sheet").strip("_") or "Sheet"


def _cell_to_text(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    return str(v)


def convert_one(xlsx: Path, out_dir: Path) -> list[Path]:
    if not xlsx.is_file():
        raise SystemExit(f"输入文件不存在: {xlsx}")
    if xlsx.suffix.lower() not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        raise SystemExit(f"仅支持 xlsx/xlsm/xltx/xltm: {xlsx}")
    out_dir.mkdir(parents=True, exist_ok=True)

    written: list[Path] = []
    base = _safe_name(xlsx.stem)

    wb = load_workbook(filename=xlsx, read_only=True, data_only=False)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            out = out_dir / f"{base}__{_safe_name(sheet_name)}.csv"
            with out.open("w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(
                    f,
                    delimiter=",",
                    quotechar='"',
                    quoting=csv.QUOTE_MINIMAL,
                    lineterminator="\n",
                )
                for row in ws.iter_rows(values_only=True):
                    w.writerow([_cell_to_text(v) for v in row])
            written.append(out)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    return written


def _write_run_info(*, xlsx: Path, out_dir: Path, written: list[Path]) -> Path:
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    info_path = out_dir / f"{_safe_name(xlsx.stem)}__conversion_info__{ts}.txt"
    lines: list[str] = [
        "xlsx_to_csv conversion info",
        f"time: {datetime.now().isoformat(timespec='seconds')}",
        f"input: {xlsx}",
        f"output_dir: {out_dir}",
        f"csv_count: {len(written)}",
        "",
        "files:",
        *[f"- {p.name}" for p in written],
        "",
    ]
    info_path.write_text("\n".join(lines), encoding="utf-8")
    return info_path


def _pick_file_gui() -> Path | None:
    # Legacy: kept for backward-compat (unused by the new GUI window)
    return None


def _pick_dir_gui() -> Path | None:
    # Legacy: kept for backward-compat (unused by the new GUI window)
    return None


def _pick_file_or_dir_gui() -> tuple[Path | None, Path | None]:
    # Legacy: kept for backward-compat (unused by the new GUI window)
    return (None, None)


def _iter_xlsx_in_dir(d: Path, *, recursive: bool) -> list[Path]:
    exts = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    it = d.rglob("*") if recursive else d.iterdir()
    out: list[Path] = []
    for p in it:
        if p.is_file() and p.suffix.lower() in exts and not p.name.startswith("~$"):
            out.append(p)
    return sorted(out)


def convert_many(
    xlsx_files: list[Path],
    *,
    out_dir: Path | None,
) -> tuple[int, list[Path]]:
    total_csv = 0
    info_files: list[Path] = []
    for x in xlsx_files:
        target_dir = out_dir if out_dir is not None else x.parent.resolve()
        written = convert_one(x, target_dir)
        info_files.append(_write_run_info(xlsx=x, out_dir=target_dir, written=written))
        total_csv += len(written)
    return (total_csv, info_files)


def _run_gui_app(*, default_out_dir_file: Path) -> int:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
    except Exception:
        raise SystemExit("GUI 依赖 tkinter，但当前环境不可用。请改用命令行参数运行。")

    root = tk.Tk()
    root.title("xlsx_to_csv")
    root.geometry("560x320")

    status = tk.StringVar(value="准备就绪。")
    resolved_default_out_dir = default_out_dir_file.resolve()

    def log(msg: str) -> None:
        status.set(msg)
        txt.configure(state="normal")
        txt.insert("end", msg + "\n")
        txt.see("end")
        txt.configure(state="disabled")

    def convert_file() -> None:
        p = filedialog.askopenfilename(
            title="选择要转换的 XLSX 文件",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"),
                ("All files", "*.*"),
            ],
        )
        if not p:
            return
        xlsx = Path(p).expanduser().resolve()
        if not xlsx.is_file():
            messagebox.showerror("错误", f"文件不存在：{xlsx}")
            return
        out_dir = resolved_default_out_dir
        # Finder/Explorer double-click sometimes yields a non-writable CWD; fall back safely.
        try:
            out_dir.mkdir(parents=True, exist_ok=True)
            probe = out_dir / ".xlsx_to_csv_write_test"
            probe.write_text("ok", encoding="utf-8")
            probe.unlink(missing_ok=True)
        except Exception:
            out_dir = xlsx.parent.resolve()
        try:
            written = convert_one(xlsx, out_dir)
            info_path = _write_run_info(xlsx=xlsx, out_dir=out_dir, written=written)
        except Exception as e:
            messagebox.showerror("转换失败", str(e))
            log(f"FAILED: {xlsx.name} -> {out_dir} ({e})")
            return
        log(f"OK: {xlsx.name} -> {out_dir} ({len(written)} CSV), info={info_path.name}")

    def convert_folder() -> None:
        p = filedialog.askdirectory(title="选择包含 XLSX 的文件夹")
        if not p:
            return
        d = Path(p).expanduser().resolve()
        if not d.is_dir():
            messagebox.showerror("错误", f"文件夹不存在：{d}")
            return

        xlsx_files = _iter_xlsx_in_dir(d, recursive=False)
        if not xlsx_files:
            messagebox.showinfo("未找到文件", f"该文件夹内未找到 xlsx：{d}")
            log(f"INFO: no xlsx in {d}")
            return

        try:
            total_csv, info_files = convert_many(xlsx_files, out_dir=d)
        except Exception as e:
            messagebox.showerror("转换失败", str(e))
            log(f"FAILED: folder {d} ({e})")
            return

        log(f"OK: folder {d} -> {d} ({len(xlsx_files)} xlsx, {total_csv} CSV)")
        for info in info_files:
            log(f"  info: {info.name}")

    frm = tk.Frame(root)
    frm.pack(fill="x", padx=12, pady=12)

    tk.Label(frm, text="选择转换方式（可多次转换，关闭窗口即可退出）").pack(anchor="w")

    btns = tk.Frame(frm)
    btns.pack(fill="x", pady=(10, 0))
    tk.Button(btns, text="选择文件并转换", command=convert_file, width=18).pack(
        side="left"
    )
    tk.Button(btns, text="选择文件夹并转换", command=convert_folder, width=18).pack(
        side="left", padx=(10, 0)
    )
    tk.Button(btns, text="关闭", command=root.destroy, width=10).pack(
        side="right"
    )

    tk.Label(
        root,
        text=f"文件模式默认输出目录：{resolved_default_out_dir}（不可写则自动回退到输入文件同目录）",
        anchor="w",
    ).pack(fill="x", padx=12)

    txt = tk.Text(root, height=10, wrap="word")
    txt.pack(fill="both", expand=True, padx=12, pady=(8, 0))
    txt.configure(state="disabled")

    bar = tk.Label(root, textvariable=status, anchor="w")
    bar.pack(fill="x", padx=12, pady=8)

    log("准备就绪。")
    root.mainloop()
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="XLSX -> UTF-8 comma CSV (per sheet)")
    ap.add_argument("--input", required=False, help="Path to .xlsx file")
    ap.add_argument("--input-dir", required=False, help="Convert all xlsx files in a folder")
    ap.add_argument(
        "--recursive",
        action="store_true",
        help="When using --input-dir, also scan subfolders",
    )
    ap.add_argument(
        "--out-dir",
        required=False,
        help="Output directory for CSV files (default: same directory as each input file)",
    )
    ap.add_argument(
        "--gui",
        action="store_true",
        help="Use GUI picker when available",
    )
    args = ap.parse_args()

    out_dir: Path | None = Path(args.out_dir).expanduser().resolve() if args.out_dir else None
    default_out_dir_file = out_dir if out_dir is not None else Path.cwd().resolve()

    # CLI: single file
    if args.input and not args.gui:
        xlsx = Path(args.input).expanduser().resolve()
        target_dir = out_dir if out_dir is not None else xlsx.parent.resolve()
        paths = convert_one(xlsx, target_dir)
        info_path = _write_run_info(xlsx=xlsx, out_dir=target_dir, written=paths)
        print(f"OK: wrote {len(paths)} CSV file(s) to {target_dir}")
        for p in paths:
            print(f" - {p.name}")
        print(f"info: {info_path.name}")
        return 0

    # CLI: folder
    if args.input_dir and not args.gui:
        d = Path(args.input_dir).expanduser().resolve()
        if not d.is_dir():
            raise SystemExit(f"输入文件夹不存在: {d}")
        xlsx_files = _iter_xlsx_in_dir(d, recursive=bool(args.recursive))
        if not xlsx_files:
            raise SystemExit(f"文件夹内未找到 xlsx: {d}")
        total_csv, info_files = convert_many(xlsx_files, out_dir=out_dir)
        print(f"OK: converted {len(xlsx_files)} xlsx file(s), wrote {total_csv} CSV file(s)")
        if out_dir is not None:
            print(f"output_dir: {out_dir}")
        for p in info_files:
            print(f"info: {p}")
        return 0

    # GUI: always open an interactive window, allow multiple rounds
    return _run_gui_app(default_out_dir_file=default_out_dir_file)


if __name__ == "__main__":
    raise SystemExit(main())

